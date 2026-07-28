import streamlit as st
import pandas as pd
import openpyxl
import io
import zipfile
import csv
from collections import defaultdict
from datetime import datetime
from openpyxl.styles import PatternFill

NULL_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

st.set_page_config(page_title="Landmark Surveys - Report Generator", layout="wide")
st.title("Track Monitoring Report Generator")
st.caption("Upload a CSV export, review the point mapping, and download the filled reports.")

TEMPLATE_PATH = "TRACK_MONITORING_REPORT.xlsx"
DEFAULT_MAPPING_PATH = "point_mapping.csv"
HEADER_ROWS = [(2, 3), (45, 46), (88, 89), (131, 132)]

# ---------- Mapping table (editable in-browser) ----------
if "mapping_df" not in st.session_state:
    st.session_state.mapping_df = pd.read_csv(DEFAULT_MAPPING_PATH)

st.subheader("1. Point mapping")
st.write("Add, remove, or edit rows below. Each Point Name needs a TMP number and the exact cells to write into.")
edited = st.data_editor(
    st.session_state.mapping_df,
    num_rows="dynamic",
    use_container_width=True,
    key="mapping_editor",
)
st.session_state.mapping_df = edited

col1, col2 = st.columns(2)
with col1:
    st.download_button(
        "Download current mapping as CSV",
        edited.to_csv(index=False),
        file_name="point_mapping.csv",
        mime="text/csv",
    )
with col2:
    uploaded_mapping = st.file_uploader("Or upload a saved mapping CSV", type="csv", key="mapping_upload")
    if uploaded_mapping:
        st.session_state.mapping_df = pd.read_csv(uploaded_mapping)
        st.rerun()

st.divider()

# ---------- CSV upload and processing ----------
st.subheader("2. Upload CSV export")
csv_file = st.file_uploader("CSV file from the field device", type="csv", key="data_csv")

def load_rounds(file):
    text = io.TextIOWrapper(file, encoding="utf-8")
    rounds = defaultdict(list)
    for row in csv.DictReader(text):
        ts = row.get("Event Time (Eastern Standard Time)")
        if ts:
            rounds[ts].append(row)
    return rounds

def safe_write(ws, cell_ref, value, fill=None):
    if not isinstance(cell_ref, str) or not CELL_PATTERN.match(cell_ref.strip()):
        return False
    ws[cell_ref.strip()] = value
    if fill:
        ws[cell_ref.strip()].fill = fill
    return True

def fill_template(rows, mapping):
    with open(TEMPLATE_PATH, "rb") as f:
        wb = openpyxl.load_workbook(f)
    ws = wb["Sheet1"]

    ts = rows[0]["Event Time (Eastern Standard Time)"]
    dt = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")
    date_str = dt.strftime("%m-%d-%y")
    time_str = dt.strftime("%I:%M%p").lstrip("0").lower()

    for date_row, time_row in HEADER_ROWS:
        ws[f"I{date_row}"] = date_str
        ws[f"I{time_row}"] = time_str

    matched, unmatched = 0, 0
    for row in rows:
        name = row["Point Name"]
        cfg = mapping.get(name)
        if not cfg:
            unmatched += 1
            continue
        ok = (safe_write(ws, cfg["Cell_DN"], float(row["StdDevNorthing"]))
              and safe_write(ws, cfg["Cell_DE"], float(row["StdDevEasting"]))
              and safe_write(ws, cfg["Cell_DELV"], float(row["StdDevElevation"])))
        if ok:
            matched += 1
        else:
            unmatched += 1

    # Any mapped point with no data this round gets NULL, highlighted yellow
    found_names = {row["Point Name"] for row in rows}
    missing_points = []
    for name, cfg in mapping.items():
        if name not in found_names:
            safe_write(ws, cfg["Cell_DN"], "NULL", NULL_FILL)
            safe_write(ws, cfg["Cell_DE"], "NULL", NULL_FILL)
            safe_write(ws, cfg["Cell_DELV"], "NULL", NULL_FILL)
            missing_points.append(name)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = dt.strftime("%m-%d-%y_%I%M%p").lstrip("0").lower() + ".xlsx"
    return fname, buf, matched, unmatched, date_str, time_str, missing_points

import re

CELL_PATTERN = re.compile(r"^[A-Z]+[0-9]+$")

def validate_mapping(mapping_df):
    """Returns (clean_lookup_dict, list_of_problem_rows)."""
    clean = {}
    problems = []
    for row in mapping_df.to_dict("records"):
        name = str(row.get("PointName", "")).strip()
        cells = [row.get("Cell_DN"), row.get("Cell_DE"), row.get("Cell_DELV")]
        if not name or name.lower() == "nan":
            continue  # blank row, ignore silently
        bad_cells = [c for c in cells if not isinstance(c, str) or not CELL_PATTERN.match(str(c).strip())]
        if bad_cells:
            problems.append(name)
            continue
        row["Cell_DN"], row["Cell_DE"], row["Cell_DELV"] = [str(c).strip() for c in cells]
        clean[name] = row
    return clean, problems

if csv_file is not None:
    mapping_lookup, problem_rows = validate_mapping(st.session_state.mapping_df)
    if problem_rows:
        st.warning(
            f"These mapping rows are missing a valid cell reference (like C15) and will be "
            f"skipped: {', '.join(problem_rows)}. Fix them in the table above and re-upload if needed."
        )
    rounds = load_rounds(csv_file)
    st.write(f"Found **{len(rounds)}** rounds in this file.")

    max_preview = min(len(rounds), 50)
    limit = st.number_input(
        "How many rounds to process (leave as full count for the whole file)",
        min_value=1, max_value=len(rounds), value=len(rounds),
    )

    if st.button("Generate reports", type="primary"):
        items = list(rounds.items())[:limit]
        zip_buf = io.BytesIO()
        total_matched, total_unmatched = 0, 0
        summary_rows = []
        progress = st.progress(0)
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for i, (ts, rows) in enumerate(items):
                fname, filebuf, matched, unmatched, date_str, time_str, missing_points = fill_template(rows, mapping_lookup)
                zf.writestr(fname, filebuf.read())
                total_matched += matched
                total_unmatched += unmatched
                summary_rows.append({
                    "Date": date_str, "Time": time_str,
                    "Points Found": matched, "Points Missing": unmatched,
                    "Missing Point Names": ", ".join(missing_points),
                    "Filename": fname,
                })
                progress.progress((i + 1) / len(items))

            summary_df = pd.DataFrame(summary_rows)
            zf.writestr("_summary_log.csv", summary_df.to_csv(index=False))
        zip_buf.seek(0)

        st.success(f"Generated {len(items)} report(s). {total_matched} point-values written, {total_unmatched} marked NULL (not found that round).")
        st.dataframe(summary_df, use_container_width=True)
        st.download_button(
            "Download all reports (.zip)",
            zip_buf,
            file_name="track_reports.zip",
            mime="application/zip",
        )
else:
    st.info("Upload a CSV to get started.")
